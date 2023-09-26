import serial
import time
import matplotlib.pyplot as plt
import numpy as np
from collections import deque
import keyboard
from scipy import signal
import openpyxl
import re

# Arduino set to COM5
# If this errors, switch the value next to COM with what COM port you are
# talking on.
ser = serial.Serial('COM5', 4800, timeout=1)

# ariable/Data structure initialization
# time.sleep(2)
num_pts = 10000
sr = 30000
data = deque()  # Pythons version of a linked list.
times = deque()
cont = True
start_time = time.perf_counter()
# prev_time = time.time()
i = 0

wb = openpyxl.Workbook()
ws = wb.active
last_row = ws.max_row + 1
ws.cell(row=1, column=1).value = "Time Elapsed"
ws.cell(row=1, column=2).value = "High Frequency Velocity (m/s)"
ws.cell(row=1, column=3).value = "Low Frequency Velocity (m/s)"

fft_of_sample = None
fft_amplitudes = None
fft_phases = None
harmonics = None


# Not the most elegant solution, but works to stop data reading for now.
def on_key_press(event):
    global cont
    if event.name == 'ctrl':
        cont = False


keyboard.on_press(on_key_press)

while cont:
    elapsed_time = time.perf_counter() - start_time
    # print(elapsed_time)

    line = ser.readline()  # read a byte
    if line:
        string = line.decode()  # convert byte string to unicode string
        string = string.replace('\r\n', '')
        if not string.isspace() or len(string) != 0:
            try:
                if "High Frequency: " in string:
                    string = re.sub("High Frequency: ", "", string)

                    # convert unicode string to int
                    num = round(float(string), 16)
                    # print(num)
                    times.append(elapsed_time)
                    data.append(num)
                    ws.cell(row=last_row+i, column=1).value = elapsed_time
                    ws.cell(row=last_row+i, column=2).value = num
                    print(num)
                elif "Low Frequency: " in string:
                    string = re.sub("Low Frequency: ", "", string)

                    # convert unicode string to int
                    num = round(float(string), 16)
                    # print(num)
                    times.append(elapsed_time)
                    data.append(num)
                    ws.cell(row=last_row+i, column=1).value = elapsed_time
                    ws.cell(row=last_row+i, column=3).value = num
                    print(num)
                i += 1
            except ValueError:
                continue  # Rejects bad string and continues while loop

keyboard.unhook_all()

wb.save('dualreadtest.xlsx')

print("Finished")

ser.close()