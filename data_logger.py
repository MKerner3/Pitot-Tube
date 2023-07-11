import math
import serial
import time
import matplotlib.pyplot as plt
import numpy as np
from collections import deque
import keyboard
from scipy import signal
import openpyxl
import nidaqmx

# Arduino set to COM5
# If this errors, switch the value next to COM with what COM port you are
# talking on.
ser = serial.Serial('COM5', 9600, timeout=1)

# ariable/Data structure initialization
# time.sleep(2)
rho = 1.15
num_pts = 10000
sr = 30000
data = deque()  # Pythons version of a linked list.
times = deque()
cont = True
start_time = time.perf_counter()
# prev_time = time.time()
i = 0

wb = openpyxl.Workbook()
ws = wb.active
last_row = ws.max_row + 1
ws.cell(row=1, column=1).value = "Time Elapsed"
ws.cell(row=1, column=2).value = "Velocity (m/s)"
ws.cell(row=1, column=3).value = "Tunnel Probe velocity (m/s)"

fft_of_sample = None
fft_amplitudes = None
fft_phases = None
harmonics = None


# Not the most elegant solution, but works to stop data reading for now.
def on_key_press(event):
    global cont
    if event.name == 'ctrl':
        cont = False


keyboard.on_press(on_key_press)

with nidaqmx.Task() as task:
    task.ai_channels.add_ai_voltage_chan("Dev1/ai6")
    while cont:
        elapsed_time = time.perf_counter() - start_time
        # print(elapsed_time)

        line = ser.readline()  # read a byte
        if line:
            if not line.isspace() or len(line) != 0:
                try:
                    # convert byte string to unicode string
                    string = line.decode()
                    string = string.replace('\r\n', '')
                    # convert unicode string to int
                    num = round(float(string), 16)
                    # rint(num)
                    times.append(elapsed_time)
                    data.append(num)

                    # wind tunnel probe voltage
                    voltage = task.read()
                    mmHg = 1.016 * voltage
                    Pa = mmHg * 133.322
                    windspeed = math.sqrt((2*Pa)/rho)
                    # print(windspeed)
                    
                    print("Test probe: " + str(num) + "| Tunnel probe: "
                          + str(windspeed))

                    ws.cell(row=last_row+i, column=1).value = elapsed_time
                    ws.cell(row=last_row+i, column=2).value = num
                    ws.cell(row=last_row+i, column=3).value = windspeed
                    # print(num)

                    i += 1
                except UnicodeDecodeError:
                    print('some weird stuff is happening')
                    break
                except ValueError:
                    continue  # Rejects bad string and continues while loop   

keyboard.unhook_all()

wb.save('graphed.xlsx')

print("Finished")

ser.close()
