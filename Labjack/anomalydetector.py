import sys
import time
import openpyxl
from labjack import ljm
import pyqtgraph as pg
from PyQt5 import QtCore, QtWidgets

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        self.xsize = 50   # Specifies the amount of data points shown
        # dynamic plot
        self.plot_graph = pg.PlotWidget()
        self.setCentralWidget(self.plot_graph)
        self.plot_graph.setBackground("w")
        pen = pg.mkPen(color=(255, 0, 0))
        #self.plot_graph.setTitle("Placeholder", color="b", size="20pt")
        styles = {"color": "red", "font-size": "18px"}
        self.plot_graph.setLabel("left", "Data", **styles)
        self.plot_graph.setLabel("bottom", "Time", **styles)
        self.plot_graph.addLegend()
        self.plot_graph.showGrid(x=True, y=True)
        self.plot_graph.setYRange(2.5,3.3)
        self.time = [0] * self.xsize

        self.ain0data = [0] * self.xsize
        self.ain1data = [0] * self.xsize
        self.ain2data = [0] * self.xsize
        self.ain3data = [0] * self.xsize

        self.previous_readings = [None, None, None, None]

        # Get a line reference
        self.line0 = self.plot_graph.plot(
            self.time,
            self.ain0data,
            name="Sensor 0",
            pen=pen,
            symbol="o",
            symbolSize=15,
            symbolBrush="y",
        )

        self.line1 = self.plot_graph.plot(
            self.time,
            self.ain1data,
            name="Sensor 1",
            pen=pen,
            symbol="x",
            symbolSize=15,
            symbolBrush="g",
        )

        self.line2 = self.plot_graph.plot(
            self.time,
            self.ain2data,
            name="Sensor 2",
            pen=pen,
            symbol="+",
            symbolSize=15,
            symbolBrush="r",
        )

        self.line3 = self.plot_graph.plot(
            self.time,
            self.ain3data,
            name="Sensor 3",
            pen=pen,
            symbol="s",
            symbolSize=15,
            symbolBrush="orange",
        )

        # Add a timer to simulate new aindata measurements
        self.check_interval = 2.0
        self.timer = QtCore.QTimer()
        self.timer.setInterval(0) # Delay between readings (in milliseconds)      ################# DELAY PARAMETER #################
        self.timer.timeout.connect(self.update_plot)
        self.timer.start()

    ################# ANOMALY THRESHOLD PARAMETERS - RADIAL AND CIRCUMFERENTIAL #################
    def check_anomalies(self, result0, result1, result2, result3, threshold = 0.2, change_threshold = 0.2):

        # Define bias for each sensor (Average voltage should be close to 2.66 V)
        result0bias = 0.01
        result1bias = 0.16
        result2bias = 0
        result3bias = 0

        # Correct bias before evaluating
        result0 = result0 - result0bias
        result1 = result1 - result1bias
        result2 = result2 - result2bias
        result3 = result3 - result3bias

        # Average of the current readings
        current_values = [result0, result1, result2, result3]
        average_value = sum(current_values) / len(current_values)
    
        # Initialize lists to store anomalies
        time_insensitive_anomalies = []
        time_sensitive_anomalies = []

        # Check for time-insensitive anomalies (between sensors)
        for i, value in enumerate(current_values):
           deviation = abs(value - average_value)
           if deviation > threshold:
                time_insensitive_anomalies.append((i + 1, value, deviation))

        # Check for time-sensitive anomalies (between time periods)
        if self.previous_readings[0] is not None:  # Ensure it's not the first run
            for i, (current, previous) in enumerate(zip(current_values, self.previous_readings)):
                change = abs(current - previous)
                if change > change_threshold:
                    time_sensitive_anomalies.append((i + 1, current, change))
    
        # Update previous readings to the current ones
        self.previous_readings = current_values

        # Return or log results
        if time_insensitive_anomalies:
            print("Time-insensitive anomalies detected (between sensors):")
            for anomaly in time_insensitive_anomalies:
                print(f"Sensor {anomaly[0]}: Value = {anomaly[1]}, Deviation = {anomaly[2]:.2f}")
    
        if time_sensitive_anomalies:
            print("Time-sensitive anomalies detected (over time):")
            for anomaly in time_sensitive_anomalies:
                print(f"Sensor {anomaly[0]}: Value = {anomaly[1]}, Change = {anomaly[2]:.2f}")
    
        # Return True if any anomaly was detected, False otherwise
        return bool(time_insensitive_anomalies or time_sensitive_anomalies)

    def update_plot(self):
        global loop_counter
        global intervalHandle
        ljm.startInterval(intervalHandle, 10000)  # Delay between readings (in microseconds)      ################# DELAY PARAMETER #################
        try:
            results = ljm.eReadNames(handle, numFrames, aNames)

            timestamp = time.time()  # Records timestamp of data log.
            elapsed = timestamp - start_time # Elapsed time (seconds)

            # Interval timer to check flow anomalies
            if (elapsed > self.check_interval):
                #self.check_anomalies(results[0], results[1], results[2], results[3])
                self.check_interval = self.check_interval + 2.0

            # prints voltages to terminal
            #print("AIN0 : %f V, AIN1 : %f V, AIN2 : %f V, AIN3 : %f V" % (results[0], results[1], results[2], results[3]))
            # Store updated values
            plot0_results.append(results[0])
            plot1_results.append(results[1])
            plot2_results.append(results[2])
            plot3_results.append(results[3])
            time_elapsed.append(elapsed)
            
            # Plot updates values on graph on graph
            self.time = self.time[1:]
            self.time.append(elapsed)

            self.ain0data = self.ain0data[1:]
            self.ain0data.append(results[0])

            self.ain1data = self.ain1data[1:]
            self.ain1data.append(results[1])

            self.ain2data = self.ain2data[1:]
            self.ain2data.append(results[2])

            self.ain3data = self.ain3data[1:]
            self.ain3data.append(results[3])

            self.line0.setData(self.time, self.ain0data)
            self.line1.setData(self.time, self.ain1data)
            self.line2.setData(self.time, self.ain2data)
            self.line3.setData(self.time, self.ain3data)
            
            ws.cell(row=last_row+loop_counter, column=1).value = elapsed
            ws.cell(row=last_row+loop_counter, column=2).value = results[0]
            ws.cell(row=last_row+loop_counter, column=3).value = results[1]
            ws.cell(row=last_row+loop_counter, column=4).value = results[2]
            ws.cell(row=last_row+loop_counter, column=5).value = results[3]

            loop_counter = loop_counter + 1

            ljm.waitForNextInterval(intervalHandle)
            if loopAmount != "infinite":
                if loop_counter >= loopAmount:
                    print("Break: Loop Amount Reached!")
                    sys.exit()
        except Exception:
            print(sys.exc_info()[1])
            print("Break: Exception!")
            sys.exit()

plot0_results = []    # Results for plot 0 at the end of the loop

plot1_results = []    # Results for plot 1 at the end of the loop

plot2_results = []    # Results for plot 2 at the end of the loop

plot3_results = []    # Results for plot 3 at the end of the loop

time_elapsed = []     # Time from first measurement to ith measurement

wb = openpyxl.Workbook()
ws = wb.active
last_row = ws.max_row + 1
ws.cell(row=1, column=1).value = "Time Elapsed (s)"
ws.cell(row=1, column=2).value = "Sensor 0 Voltage (V)"
ws.cell(row=1, column=3).value = "Sensor 1 Voltage (V)"
ws.cell(row=1, column=2).value = "Sensor 2 Voltage (V)"
ws.cell(row=1, column=3).value = "Sensor 3 Voltage (V)"

loopMessage = ""
if len(sys.argv) > 1:
    # An argument was passed. The first argument specifies how many times to
    # loop.
    try:
        loopAmount = int(sys.argv[1])
    except:
        raise Exception("Invalid first argument \"%s\". This specifies how many"
                        " times to loop and needs to be a number." %
                        str(sys.argv[1]))
else:
    # An argument was not passed. Loop an infinite amount of times.
    loopAmount = "infinite"
    loopMessage = " Exit plot window to stop."

loop_counter = 0 # loop counter
intervalHandle = 1


# Open first found LabJack
handle = ljm.openS("T7", "USB", "ANY")  # T7 device, USB connection, Any identifier

info = ljm.getHandleInfo(handle)
print("Opened a LabJack with Device type: %i, Connection type: %i,\n"
      "Serial number: %i, IP address: %s, Port: %i,\nMax bytes per MB: %i" %
      (info[0], info[1], info[2], ljm.numberToIP(info[3]), info[4], info[5]))

deviceType = info[0]

# Setup and call eWriteNames to configure AIN0-3 on the LabJack.
if deviceType == ljm.constants.dtT4:
    # LabJack T4 configuration

    # AIN0-3:
    #   Resolution index = Default (0)
    #   Settling, in microseconds = Auto (0)
    aNames = ["AIN0_RESOLUTION_INDEX", "AIN0_SETTLING_US",
              "AIN1_RESOLUTION_INDEX", "AIN1_SETTLING_US",
              "AIN2_RESOLUTION_INDEX", "AIN2_SETTLING_US",
              "AIN3_RESOLUTION_INDEX", "AIN3_SETTLING_US"]
    aValues = [0, 0, 0, 0, 0, 0, 0, 0]
else:
    # LabJack T7 and T8 configuration

    # AIN0 and AIN2:
    #   Range: +/-10.0 V (10.0)
    #   Resolution index = Default (0)
    aNames = ["AIN0_RANGE", "AIN0_RESOLUTION_INDEX",
              "AIN1_RANGE", "AIN1_RESOLUTION_INDEX",
              "AIN2_RANGE", "AIN2_RESOLUTION_INDEX",
              "AIN3_RANGE", "AIN3_RESOLUTION_INDEX"]
    aValues = [10.0, 0, 10.0, 0, 10.0, 0, 10.0, 0]

    # Negative channel and settling configurations do not apply to the T8
    if deviceType == ljm.constants.dtT7:
        #     Negative Channel = 199 (Single-ended)
        #     Settling = 0 (auto)
        aNames.extend(["AIN0_NEGATIVE_CH", "AIN0_SETTLING_US",
                       "AIN1_NEGATIVE_CH", "AIN1_SETTLING_US",
                       "AIN2_NEGATIVE_CH", "AIN2_SETTLING_US",
                       "AIN3_NEGATIVE_CH", "AIN3_SETTLING_US"])
        aValues.extend([199, 0, 199, 0, 199, 0, 199, 0])

numFrames = len(aNames)
ljm.eWriteNames(handle, numFrames, aNames, aValues)

print("\nSet configuration:")

for i in range(numFrames):
    print("    %s : %f" % (aNames[i], aValues[i]))

# Read AIN0-3 from the LabJack with eReadNames in a loop.
numFrames = 4
aNames = ["AIN0", "AIN1", "AIN2", "AIN3"]

print("\nStarting %s read loops.%s\n" % (str(loopAmount), loopMessage))
start_time = time.time()   # Records starting time of the loop.

app = QtWidgets.QApplication([])
main = MainWindow()
main.show()

# Clean up running labjack data collection and real time loop
app.exec_() 
ljm.cleanInterval(intervalHandle)
ljm.close(handle)

# Post Processing #
#wb.save('lowhightest.xlsx')